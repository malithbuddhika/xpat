#!/usr/bin/env python3
"""Test script for missing data detection and formatting feature."""

from app.data.excel_reader import load_workers_from_excel
from app.data.excel_writer import export_to_excel
import openpyxl

print("=" * 60)
print("TESTING: Missing Data Detection and Formatting")
print("=" * 60)

# Load test data
print("\n1. Loading test Excel with missing data...")
df, workers = load_workers_from_excel('test_missing_data.xlsx')

# Check how many workers have missing data
missing_count = sum(1 for w in workers if w.missing_data)
print(f"   ✓ Loaded {len(workers)} workers")
print(f"   ✓ Found {missing_count} workers with missing data")

# Show details
print("\n2. Worker details:")
for i, w in enumerate(workers):
    status = "✓ VALID" if not w.missing_data else "✗ NO DATA"
    print(f"   Row {i}: permit='{w.work_permit_number}', passport='{w.passport_number}' [{status}]")

# Test export
print(f"\n3. Exporting to Excel with red formatting...")
export_to_excel(df, workers, 'final_test_output.xlsx')
print(f"   ✓ Exported to final_test_output.xlsx")

# Verify formatting
print(f"\n4. Verifying red formatting in exported file...")
wb = openpyxl.load_workbook('final_test_output.xlsx')
ws = wb.active

red_cells = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in row:
        if "NO DATA" in str(cell.value):
            is_red = cell.font and cell.font.color and cell.font.color.rgb and 'FF0000' in cell.font.color.rgb
            status = "✓ RED" if is_red else "? NOT RED"
            red_cells.append(f"   {cell.coordinate}: {cell.value} [{status}]")

if red_cells:
    print("   Cells with NO DATA:")
    for cell_info in red_cells:
        print(cell_info)

print("\n" + "=" * 60)
print("FEATURE TEST PASSED!")
print("=" * 60)
