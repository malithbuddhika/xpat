"""Excel/CSV writing helpers."""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from ..models.worker import Worker
from ..utils.config import OUTPUT_COLUMNS
from ..utils.logger import get_logger

logger = get_logger()


def _normalize_key(value: Optional[str]) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def _map_workers_by_permit(workers: List[Worker]) -> Dict[str, Worker]:
    return {w.work_permit_number.strip().upper(): w for w in workers}


def export_to_excel(
    base_df: pd.DataFrame,
    workers: List[Worker],
    output_path: str,
    include_index: bool = False,
) -> None:
    """Export results to a new Excel file."""
    df = base_df.copy()
    mapping = _map_workers_by_permit(workers)

    for col in OUTPUT_COLUMNS:
        df[col] = ""

    # Track which cells need red formatting (for NO DATA cells)
    cells_to_format_red = set()

    for idx, row in df.iterrows():
        # Get original values (before normalization)
        original_permit = row.get("WorkPermitNumber")
        original_passport = row.get("PassportNumber")
        
        # Normalize for lookup
        key = _normalize_key(original_permit)
        
        # Mark WorkPermitNumber and PassportNumber as NO DATA if they're empty (including NaN)
        permit_is_empty = pd.isna(original_permit) or str(original_permit).strip() == ""
        passport_is_empty = pd.isna(original_passport) or str(original_passport).strip() == ""
        
        if permit_is_empty:
            df.at[idx, "WorkPermitNumber"] = "NO DATA"
            cells_to_format_red.add((idx + 2, 1))  # +2 for header and 0-indexing
        
        if passport_is_empty:
            df.at[idx, "PassportNumber"] = "NO DATA"
            cells_to_format_red.add((idx + 2, 2))  # +2 for header and 0-indexing
        
        # Skip worker data if no valid permit number
        if not key:
            continue
            
        worker = mapping.get(key)
        if not worker:
            continue
        
        # Only add worker data if not marked as NO DATA
        if not permit_is_empty and not passport_is_empty:
            df.at[idx, "WorkerName"] = worker.worker_name or ""
            df.at[idx, "Company"] = worker.company or ""
            df.at[idx, "Status"] = worker.status or ""
            df.at[idx, "IssuedDate"] = worker.issued_date or ""
            df.at[idx, "PermitValidTill"] = worker.permit_valid_till or ""
            df.at[idx, "WorkSite"] = worker.work_site or ""
            df.at[idx, "VerificationTime"] = worker.verification_time or ""
            df.at[idx, "WorkerPhotoURL"] = worker.photo_url or ""

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if output_file.suffix.lower() == ".csv":
        df.to_csv(output_file, index=include_index)
    else:
        # Write to Excel
        df.to_excel(output_file, engine="openpyxl", index=include_index)
        
        # Apply red formatting to NO DATA cells
        if cells_to_format_red:
            wb = load_workbook(output_file)
            ws = wb.active
            red_font = Font(color="FF0000", bold=True)
            
            for row_idx, col_idx in cells_to_format_red:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = red_font
            
            wb.save(output_file)

    logger.info("Exported results to %s", output_file)
