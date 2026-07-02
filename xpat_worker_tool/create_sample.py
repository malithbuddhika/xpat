#!/usr/bin/env python3
"""Create sample Excel template for XPAT Worker Tool."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Workers"

# Add headers
headers = ["Work Permit Number", "Passport Number"]
ws.append(headers)

# Style header
for col_num, header_text in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")

# Add sample rows
samples = [
    ["WP123456789", "AB987654321"],
    ["WP234567890", "CD123456789"],
    ["WP345678901", "EF234567890"],
]
for row_data in samples:
    ws.append(row_data)

# Set column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20

# Save
output = Path("assets/sample_template.xlsx")
output.parent.mkdir(exist_ok=True)
wb.save(output)
print(f"✅ Sample template created: {output}")
